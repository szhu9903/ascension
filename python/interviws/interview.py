#22、s = "ajldjlajfdljfddd"，去重并从小到大排序输出"adfjl"
'''
s = "ajldjlajfdljfddd"
s = set(s)
s = list(s)
s.sort(reverse=False)
res = ''.join(s)
print(s)
'''

#23、用lambda函数实现两个数相乘
'''
sum = lambda a,b:a*b
print(sum(4,5))
'''

#24、字典根据键从小到大排序
'''
dic = {"name":"zs","age":18,"city":"许昌","tel":"1362626627"}
lis = sorted(dic.items(),key=lambda i:i[0],reverse=False)
print(dict(lis))
'''

#26、字符串a = "not 404 found 张三 99 深圳"，每个词中间是空格，用正则过滤掉英文和数字，最终输出"张三 深圳"
'''
import re
a = "not 404 found 张三 99 深圳"
list = a.split(' ')
#p匹配非重复的正则，返回一个列表，然后删除两个列表重复的部分
res = re.findall('\d+|[a-zA-Z]+',a)
for i in res:
	if i in list:
		list.remove(i)
new_list = ' '.join(list)
print(new_list)
'''

#27、filter方法求出列表所有奇数并构造新列表，a = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10],列表推倒式
'''
a = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]
def fn(a):
	return a%2 == 1
g = filter(fn,a)
print(g)
new_list = [i for i in g]
print(new_list)

new_lists = [i for i in a if i%2 == 1]
print(new_lists)
'''

#打印现在的时间
'''
import datetime
a = str(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')) +'星期'+ str(datetime.datetime.now().isoweekday())
print(a)
'''

#39、[[1,2],[3,4],[5,6]]一行代码展开该列表，得出[1,2,3,4,5,6]
'''
a = [[1,2],[3,4],[5,6]]
new_list = [i for j in a for i in j]
print(new_list)
'''

#71、举例sort和sorted对列表排序，list=[0,-1,3,-10,5,9]
'''
list = [0,-1,3,-10,5,9]
print(id(list))
list.sort(reverse = False)
print(id(list),list)
print('----------sort在列表的基础上做出排序，sorted生成一个新的列表----------')

list = [0,-1,3,-10,5,9]
print(id(list))
# res = sorted(list,reverse=False)
res = sorted(list,key = lambda x:(x<0,abs(x)))#abs()绝对值
print(id(res),res)
'''

#74、列表嵌套字典的排序，分别根据年龄和姓名排序
'''
foo = [{"name":"zs","age":19},{"name":"ll","age":54},
{"name":"wa","age":17},{"name":"df","age":23}]

a = sorted(foo,key = lambda x:x['age'])
print(a)
a = sorted(foo,key = lambda x:x['name'])
print(a)
'''

# 78、根据键对字典排序（方法二,不用zip)
'''dic = {'name','red','man','good','nice'}
 a = sorted(dic.items(),key=lambda x:x[0])
 new_dic = {i[0]:i[1] for i in a}'''

#生成器
'''
import random

dic = {k:random.randint(5,10) for k in ['a','b','c','d']}
print(dic)
'''

#递归求和
'''
def get_num(num):
	if num>0:
		res = num+get_num(num-1)
	else:
		res = 0
	return res
print(get_num(100))
'''

#101、求两个列表的交集、差集、并集
'''
a = [1,2,3,4]
b = [3,4,5,6]

print('交集：',[i for i in a if i in b])
print('交集:',list(set(a).intersection(set(b))))

print('并集:',list(set(a).union(set(b))))

print('差集：',list(set(a).difference(set(b))))
print('差集：',list(set(b).difference(set(a))))
'''

#迭代器，适合遍历比较巨大的集合。__iter__:返回迭代器本身，__next__：返回容器中下一个元素或数据
'''
for x in range(101):
	print('生成%s'%x)
'''

#生成器，带有yield的函数，当函数被调用时返回一个生成器对象，生成器函数生成值后会自动挂起并暂停他们的执行状态
'''
def myyield(n):
	while n>=0:
		print('生成%s'%n)
		yield n
		n -= 1
print([n for n in myyield(100)])
'''

#进制转换
'''
print(bin(8))
print(oct(8))
print(hex(8))

a = {'%s is square'%i:i**2 for i in range(11)}
b = a.keys()
print(a)
print(b)
'''

#当前时间
'''
import datetime

time = datetime.datetime.now()
print(time)
'''

# StringIO
'''
from io import StringIO
output = StringIO()
print(output.write('python'))#写入
print(output.seek(0))#将指针定位到起始位置
print(output.read())#读取写入的内容
print(output.close())#close
'''

# BytesIO
'''
from io import BytesIO
output = BytesIO()
output.write(b'python')
output.seek(0)
print(output.read())
output.close()
'''

# pandas
'''
from pandas import Series,DataFrame
import pandas as pd

# s = Series([100,'python','numpy','pandas','series','dateframe'],index = [10,'py','np','pd','si','df'])
data = {"name":['java','c#','python','javascript'],"marks":[200,400,600,800],"price":[9,3,7,10]}
f = DataFrame(data)
print(f)
'''

# Python中的模块--XlsxWriter
'''
import xlsxwriter

#创建一个工作簿添加一张工作表
workbook = xlsxwriter.Workbook('cookbook.xlsx')
worksheet = workbook.add_worksheet()

#要插入的数据
expenses = (
	['python',1000],
	['java',500],
	['C++',6000],
	['Go',500]
)

#从第一个单元格开始，行和列索引均为0
row = 0
col = 0

#迭代数据并逐行写入
for item,cost in (expenses):
	worksheet.write(row,col,item)
	worksheet.write(row,col+1,cost)
	row +=1

#计算总和
worksheet.write(row,0,'Total')
worksheet.write(row,1,'=SUM(B1:B4)')

workbook.close()
'''

# padas 生成excel 增加sheet表
'''
import sys,importlib
import pandas as pd 
from openpyxl import load_workbook

importlib.reload(sys)

# pandas dataframe 生成excel表
def dataFrame2sheet(dataframe,excelwrite):
    # dataframe转换为excel中的sheet
	dataframe.to_excel(excel_writer=excelwrite,sheet_name='zhu1',index=None)
	dataframe.to_excel(excel_writer=excelwrite,sheet_name='zhu2',index=None)
	dataframe.to_excel(excel_writer=excelwrite,sheet_name='zhu3',index=None)
	dataframe.to_excel(excel_writer=excelwrite,sheet_name='zhu4',index=None)
	excelwrite.save()
	excelwrite.close()

# excel中添加sheet表
def excelAddSheet(dataframe,excelwrite):
	book = load_workbook(excelwrite.path)
	excelwrite.book = book
	dataframe.to_excel(excel_writer=excelwrite,sheet_name='zhu5',index=None)
	excelwrite.close()


if __name__ == "__main__":
	# 数据集
	dateset = [
		{'姓名':'朱','年龄':22,'性别':'男'},
		{'姓名':'热','年龄':22,'性别':'男'},
		{'姓名':'各','年龄':22,'性别':'女'},
		{'姓名':'宁','年龄':22,'性别':'女'}
		]
	
	#excelPath
	excelPath = "C:\\Users\Administrator\Desktop\python提升\zhu.xlsx"

	#生成DataFrame
	dataframe = pd.DataFrame(dateset)

	#创建ExcelWrite 对象
	excelwrite = pd.ExcelWriter(excelPath,engine = 'openpyxl')

	#生成excel
	dataFrame2sheet(dataframe,excelwrite)

	# excel中添加sheet
	excelAddSheet(dataframe,excelwrite)
'''

# 读取数据
"""
import pandas as pd

# 读取前n行所有数据
excelpath = "F:\测试模板\费用模板"
df = pd.read_excel('%s\住宿费.xlsx'%excelpath)
data1 = df.head(7)#读取前七行所有数据
data2 = df.values #list形式读取表中所有数据
print('{0}\n'.format(data1))
print('{0}\n'.format(data2))

# 读取特定行、特定列
data3 = df.ix[0].values #读取第一行的所有数据
data4 = df.ix[1,2] #读取指定位置数据
data5 = df.ix[[1,2]].values #读取指定多行
data6 = df.ix[:,[0]].values #指定列的的所有行
data7 = df.ix[:,[u'房间',u'备注']].values  #指定列的的所有行
print('{0}\n'.format(data3))
print('{0}\n'.format(data4))
print('{0}\n'.format(data5))
print('{0}\n'.format(data6))
print('{0}\n'.format(data7))

# 获取xlsx文件行号，所有列名称
print('输出行号列表%s'%df.index.values)
print('输出列标题%s'%df.columns.values)

#读取xlsx数据转换为字典
test_data = []
for i in df.index.values:
	#用i来获取每一行指定的数据，利用to_dict转为字典
	row_data = df.ix[i,['序号','房间','开始日期','结束日期','备注']].to_dict()
	test_data.append(row_data)
print('最终取到的数据是：%s'%test_data)
"""

# 写入xslx文件
"""
import xlsxwriter
# 创建工作簿
file_name = "小朱.xlsx"
workbook = xlsxwriter.Workbook(file_name)
#创建工作表
worksheet = workbook.add_worksheet('sheet1')
#写入单元格
worksheet.write(0,0,'序号')
worksheet.write(0,1,'姓名')
worksheet.write(0,2,'年龄')
#行
worksheet.write_row(1,0,[1,'朱帅杰',22])
#列
worksheet.write_column('A3',[i for i in range(11) if i>1])
#关闭工作簿
workbook.close()
"""

# xpython中enumerate()函数的用法
"""
import json
def e_format(e_key,e_value):
    rs = []
    for v in e_value:
        dic = {}
        for index,values in enumerate(e_key):
            dic[values] = v[index]
        rs.append(dic)
    return rs

if __name__ == '__main__':
    e_key = ['id','name','age']
    e_value = [(2,'朱',22),(3,'朱',22),(4,'朱',22)]
    rs = e_format(e_key,e_value)
    print(json.dumps(rs))
"""









































